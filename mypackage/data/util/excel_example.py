import openpyxl


# 写入
def write07Excel(path):
	#创建文件
    wb = openpyxl.Workbook()
    #创建sheet，修改sheet的名字
    sheet = wb.active
    sheet.sheet_properties.tabColor = "1072BA"
    sheet.title = '2007测试表'
    #在末尾创建另一张sheet
    sheet2 = wb.create_sheet("Mysheet")

    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    # 保存文件在path地址
    wb.save(path)
    print("写入数据成功！")
    print('----------')

# 读取
def read07Excel(path):
	#打开并读取文件
    wb = openpyxl.load_workbook(path)
    #读取指定sheet
    sheet = wb.get_sheet_by_name('2007测试表')

    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()
    print('读取完成')


file_2007 = '/Users/jk/www/interface/mypackage/data_config2007.xlsx'
write07Excel(file_2007)
read07Excel(file_2007)