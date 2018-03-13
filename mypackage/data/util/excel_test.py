import openpyxl

def readexcel(path):
	wb = openpyxl.load_workbook(path)
	sheet = wb.get_sheet_by_name('2007测试表')
	#不知道名字用index
    # = wb.get_sheet_by_name(sheet_names[index])# index为0为第一张表 
    
	# 获取行数
	rows=sheet.max_row
	print('行数是：%d'%rows)
	# 获取列数
	column=sheet.max_column
	print("列数是：%d"%column)
	# Data=table.cell(row=row,column=col).value  #获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
	Data=sheet.cell(row=2,column=2).value
	print(Data)

file_2007 = '/Users/jk/www/ImoocInterace/data/2007.xlsx'
readexcel(file_2007)