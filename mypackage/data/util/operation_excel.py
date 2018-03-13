import openpyxl


class OperationExcel(object):
    def __init__(self, path=None, sheet_name=None):
        if path:
            self.path = path
            self.sheet_name = sheet_name
        else:
            self.path = '/Users/jk/www/interface/mypackage/data_config/mycase.xlsx'
            self.sheet_name = 'sheet1'
        self.data = self.get_data()

    # 获取sheet内容
    def get_data(self):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(self.sheet_name)
        return sheet

    # 获取行数
    def get_rows(self):
        rows = self.data.max_row
        return rows

    # 获取列数
    def get_column(self):
        column = self.data.max_column
        return column

    # 获得单元格内容
    def get_value(self, row_num, colum_num):
        value = self.data.cell(row=row_num, column=colum_num).value
        return value

    # 单元格写入数据
    def revise_value(self, row, col, value):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(self.sheet_name)
        sheet.cell(row=row, column=col, value=value)
        # 保存
        wb.save('/Users/jk/www/interface/mypackage/data_config/mycase.xlsx')

    # 根据case_id找到对应行的内容
    def get_rows_data(self, case_id):
        row_num = self.get_rows_num(case_id)
        print(row_num)
        rows_data = self.get_rows_value(row_num)
        return rows_data

    # 根据对应的case_id找到对应的行号
    def get_rows_num(self, case_id):
        num = 0
        data = self.get_column_value()
        for col_data in data:
            if case_id in col_data:
                return num
            num = num + 1

    #根据行号，找到该行的内容
    def get_rows_value(self, row):
        data = []
        tables = self.data
        # tables.rows时生成器类型无法序列化，转换为list类型
        for cell in list(tables.rows)[row]:
            data.append(cell.value)
        return data

    # 获取某一列的内容
    def get_column_value(self, col_id=None):
        data = []
        tables = self.data
        if col_id:
            for cell in list(tables.columns)[col_id]:
                data.append(cell.value)
        else:
            for cell in list(tables.columns)[0]:
                data.append(cell.value)
        return data

if __name__ == '__main__':
#     excel_path = ''
#     excel_sheet_name = ''
# #     # 实例化对象
#     opens = OperationExcel(excel_path,excel_sheet_name)
      opens = OperationExcel()
      print(opens.get_rows_value(1))
      print(opens.get_column_value())
      print(opens.get_rows_data('jk-1'))
#     values = opens.get_value(2,12)
#     print(values)
#     opens.revise_value(2,12,'通过')
#     values = opens.get_value(2,12)
# #     lines = opens.get_rows()
# #     cols = opens.get_column()
# #     print(cols)
# #     print(lines)
#     print(values)
